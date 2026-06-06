"""AI Report Agent — FastAPI 後端。

把原 customtkinter 桌面 app 的核心邏輯(app/ 套件)包成本機 HTTP/WS 服務,
供 Vite + React 前端呼叫。單一使用者、本機執行,因此用一份全域狀態即可。
"""

import json
import os
import queue
import threading
import time
from pathlib import Path

# 載入 API key(GEMINI_API_KEY 等):優先本專案 .env,其次沿用 AI- 專案的 .env
try:
    from dotenv import load_dotenv
    for _envp in (Path(__file__).parent / ".env",
                  Path(__file__).parent.parent / ".env",
                  Path(__file__).parent.parent.parent / "AI-" / ".env"):
        if _envp.exists():
            load_dotenv(_envp)
except Exception:
    pass

import pandas as pd
import uvicorn
from fastapi import FastAPI, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from app.settings import load_settings, save_settings, DEFAULTS
from app.agent.budget import BudgetTracker
from headless_context import HeadlessContext
import native_dialog

# ------------------------------------------------------------
# 全域狀態(單機單使用者)
# ------------------------------------------------------------
_settings = load_settings()
_budget = BudgetTracker(
    planner_limit=int(_settings.get("max_planner_calls", 50)),
    reviewer_limit=int(_settings.get("max_reviewer_calls", 100)),
)
ctx = HeadlessContext(settings=_settings, budget=_budget)

app = FastAPI(title="AI Report Agent")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_methods=["*"], allow_headers=["*"],
)


# index.html / 首頁永不快取 → 前端更新後瀏覽器一定拿到最新版(避免一直看到舊畫面)。
# 內容雜湊過的 /assets/*.js|css 可長快取(檔名變了就自動換)。
@app.middleware("http")
async def _no_cache_html(request, call_next):
    resp = await call_next(request)
    path = request.url.path
    if path == "/" or path.endswith(".html") or path == "/index.html":
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
    return resp


# ------------------------------------------------------------
# 設定
# ------------------------------------------------------------
@app.get("/api/settings")
def get_settings():
    return ctx.settings


class SettingsPatch(BaseModel):
    patch: dict


@app.post("/api/settings")
def update_settings(body: SettingsPatch):
    for k, v in body.patch.items():
        if k in DEFAULTS:
            ctx.settings[k] = v
    save_settings(ctx.settings)
    # 同步預算上限
    _budget.update_limits(
        planner_limit=int(ctx.settings.get("max_planner_calls", 50)),
        reviewer_limit=int(ctx.settings.get("max_reviewer_calls", 100)),
    )
    return {"ok": True, "settings": ctx.settings}


# ------------------------------------------------------------
# 原生檔案選取
# ------------------------------------------------------------
class PickBody(BaseModel):
    kind: str = "any"


@app.post("/api/pick")
def pick_file(body: PickBody):
    path = native_dialog.pick(body.kind)
    return {"path": path, "cancelled": not path}


# ------------------------------------------------------------
# Excel / 範本 讀取
# ------------------------------------------------------------
class PathBody(BaseModel):
    path: str
    sheet: str = ""
    header_row: int = 1
    preview_rows: int = 8


@app.post("/api/excel/sheets")
def excel_sheets(body: PathBody):
    if not body.path or not os.path.isfile(body.path):
        return {"error": "檔案不存在", "sheets": []}
    try:
        return {"sheets": pd.ExcelFile(body.path).sheet_names}
    except Exception as e:
        return {"error": str(e), "sheets": []}


@app.post("/api/excel/columns")
def excel_columns(body: PathBody):
    if not body.path or not os.path.isfile(body.path):
        return {"error": "檔案不存在", "columns": [], "rows": []}
    try:
        df = pd.read_excel(
            body.path,
            sheet_name=body.sheet if body.sheet else 0,
            header=max(0, body.header_row - 1),
        )
        cols = [str(c) for c in df.columns]
        head = df.head(body.preview_rows).fillna("")
        rows = [[str(v) for v in row] for row in head.values.tolist()]
        return {"columns": cols, "rows": rows, "total_rows": int(len(df))}
    except Exception as e:
        return {"error": str(e), "columns": [], "rows": []}


@app.post("/api/template/word-vars")
def word_vars(body: PathBody):
    if not body.path or not os.path.isfile(body.path):
        return {"error": "檔案不存在", "variables": []}
    try:
        from docxtpl import DocxTemplate
        v = sorted(DocxTemplate(body.path).get_undeclared_template_variables())
        return {"variables": v}
    except Exception as e:
        return {"error": str(e), "variables": []}


@app.post("/api/template/word-text")
def word_text(body: PathBody):
    """讀 docx 段落文字,供視覺化標籤編輯器顯示與插入錨點。"""
    from app.agent.template_edit import read_docx_text
    return read_docx_text(body.path, max_paragraphs=0)


@app.post("/api/template/excel-vars")
def excel_template_vars(body: PathBody):
    from app.excel_template import ExcelReportGenerator
    if not body.path or not os.path.isfile(body.path):
        return {"error": "檔案不存在", "variables": []}
    try:
        gen = ExcelReportGenerator(template_path=body.path, excel_path="",
                                   sheet_name=body.sheet)
        return {"variables": sorted(gen.template_variables())}
    except Exception as e:
        return {"error": str(e), "variables": []}


@app.post("/api/template/excel-grid")
def excel_grid(body: PathBody):
    """讀 Excel 範本指定 sheet 的儲存格內容(供視覺化網格編輯器)。"""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    if not body.path or not os.path.isfile(body.path):
        return {"error": "檔案不存在", "cells": []}
    try:
        wb = load_workbook(body.path, data_only=False)
        ws = wb[body.sheet] if body.sheet and body.sheet in wb.sheetnames else wb.active
        max_r = min(ws.max_row or 1, 60)
        max_c = min(ws.max_column or 1, 20)
        cells = []
        for r in range(1, max_r + 1):
            row = []
            for c in range(1, max_c + 1):
                v = ws.cell(row=r, column=c).value
                row.append("" if v is None else str(v))
            cells.append(row)
        cols = [get_column_letter(c) for c in range(1, max_c + 1)]
        return {"cells": cells, "col_labels": cols, "sheet": ws.title,
                "sheets": wb.sheetnames}
    except Exception as e:
        return {"error": str(e), "cells": []}


# ------------------------------------------------------------
# 驗證
# ------------------------------------------------------------
class ValidateBody(BaseModel):
    mode: str = "word"  # word | excel


@app.post("/api/template/pptx-vars")
def pptx_vars(body: PathBody):
    from app.pptx_template import PptxReportGenerator
    if not body.path or not os.path.isfile(body.path):
        return {"error": "檔案不存在", "variables": []}
    try:
        g = PptxReportGenerator(template_path=body.path, excel_path="")
        return {"variables": sorted(g.template_variables())}
    except Exception as e:
        return {"error": str(e), "variables": []}


@app.post("/api/validate")
def validate(body: ValidateBody):
    if body.mode == "excel":
        return ctx.validate_excel_template()
    if body.mode == "pptx":
        return ctx.validate_pptx_template()
    return ctx.validate_template()


# ------------------------------------------------------------
# 範本標籤編輯
# ------------------------------------------------------------
class InsertTagBody(BaseModel):
    anchor: str
    variable: str
    position: str = "after"


@app.post("/api/template/insert-tag")
def insert_tag(body: InsertTagBody):
    return ctx.insert_template_variable(body.anchor, body.variable, body.position)


class RenameTagBody(BaseModel):
    old: str
    new: str


@app.post("/api/template/rename-tag")
def rename_tag(body: RenameTagBody):
    return ctx.rename_template_variable(body.old, body.new)


class ExcelInsertTagBody(BaseModel):
    cell: str
    tag: str


@app.post("/api/template/excel-insert-tag")
def excel_insert_tag(body: ExcelInsertTagBody):
    from app.excel_template import insert_tag_into_excel_cell
    path = ctx._g("excel_template_path")
    sheet = ctx._g("excel_template_sheet")
    if not path:
        return {"error": "未設定 Excel 範本路徑"}
    try:
        coord = insert_tag_into_excel_cell(path, sheet, body.cell, body.tag)
        return {"ok": True, "cell": coord}
    except Exception as e:
        return {"error": str(e)}


@app.post("/api/template/suggest-mappings")
def suggest_mappings():
    return ctx.suggest_mappings()


# ------------------------------------------------------------
# 資料夾圖片 → 欄位 對應
# ------------------------------------------------------------
class ImgFolderBody(BaseModel):
    folder: str


@app.post("/api/images/list")
def images_list(body: ImgFolderBody):
    from app.image_mapper import list_images
    return {"images": list_images(body.folder)}


class ImgMatchBody(BaseModel):
    image_names: list
    targets: list
    use_ai: bool = False
    hint: str = ""


@app.post("/api/images/match")
def images_match(body: ImgMatchBody):
    from app.image_mapper import deterministic_match, ai_match
    if body.use_ai:
        llm, model, err = ctx._build_planner_client()
        if err:
            return {"error": err, "mapping": deterministic_match(body.image_names, body.targets)}
        if _budget.can_use_planner():
            res = ai_match(llm, model, body.image_names, body.targets, body.hint)
            _budget.use_planner()
            if "error" in res:
                return {"error": res["error"], "mapping": deterministic_match(body.image_names, body.targets)}
            return {"mapping": res, "ai": True}
        return {"error": "已達 planner 預算上限", "mapping": deterministic_match(body.image_names, body.targets)}
    return {"mapping": deterministic_match(body.image_names, body.targets)}


class ImgApplyStaticBody(BaseModel):
    mapping: dict  # {field: image_path}
    width_mm: int = 80


@app.post("/api/images/apply-static")
def images_apply_static(body: ImgApplyStaticBody):
    from app.image_mapper import apply_static_to_template
    wp = ctx._g("word_path")
    if not wp:
        return {"error": "未設定 Word 範本路徑"}
    return apply_static_to_template(wp, body.mapping, width_mm=body.width_mm)


class ImgVisualBody(BaseModel):
    folder: str
    width_mm: int = 0
    use_com: bool = True


@app.post("/api/images/auto-visual")
def images_auto_visual(body: ImgVisualBody):
    return ctx.auto_place_images_visual(body.folder, width_mm=body.width_mm, use_com=body.use_com)


class FillVisualBody(BaseModel):
    word_path: str = ""
    row_index: int = 0


@app.post("/api/text/auto-visual")
def text_auto_visual(body: FillVisualBody):
    return ctx.auto_fill_text_visual(word_path=body.word_path, row_index=body.row_index)


class ReportFillBody(BaseModel):
    template_path: str
    photo_root: str
    output_path: str = ""
    dry_run: bool = False
    match_mode: str = "auto"


@app.post("/api/report/fill-from-folders")
def report_fill_from_folders(body: ReportFillBody):
    return ctx.fill_report_from_folders(
        body.template_path, body.photo_root,
        output_path=body.output_path, dry_run=body.dry_run, match_mode=body.match_mode)


class ImgFillExcelBody(BaseModel):
    folder: str
    key_column: str
    image_column: str
    image_to_key: dict = None


@app.post("/api/images/fill-excel")
def images_fill_excel(body: ImgFillExcelBody):
    from app.image_mapper import fill_excel_image_column
    ep = ctx._g("excel_path")
    if not ep:
        return {"error": "未設定來源 Excel"}
    res = fill_excel_image_column(
        ep, ctx._g("sheet_name"), ctx._int("header_row", 1),
        body.key_column, body.image_column, body.folder,
        image_to_key=body.image_to_key,
    )
    # 成功則把資料來源切到含圖片路徑的副本,讓產出自動嵌圖
    if res.get("output_path"):
        ctx._set("excel_path", res["output_path"])
    return res


# ------------------------------------------------------------
# 產出(SSE 串流進度)
# ------------------------------------------------------------
def _sse(event: dict) -> str:
    return f"data: {json.dumps(event, ensure_ascii=False, default=str)}\n\n"


def _run_generation_stream(mode: str):
    q = queue.Queue()

    def progress(produced, total, saved_path, row_dict):
        q.put({"type": "progress", "produced": produced, "total": total,
               "file": os.path.basename(str(saved_path))})

    def worker():
        try:
            if mode == "excel":
                result = ctx.generate_excel_reports(progress_callback=progress)
            elif mode == "pptx":
                result = ctx.generate_pptx_reports(progress_callback=progress)
            else:
                result = ctx.generate_reports(progress_callback=progress)
            q.put({"type": "done", "result": result})
        except Exception as e:
            q.put({"type": "error", "error": str(e)})
        finally:
            q.put(None)

    threading.Thread(target=worker, daemon=True).start()

    yield _sse({"type": "start", "mode": mode})
    while True:
        item = q.get()
        if item is None:
            break
        yield _sse(item)


@app.get("/api/generate/stream")
def generate_stream(mode: str = "word"):
    return StreamingResponse(_run_generation_stream(mode),
                             media_type="text/event-stream")


@app.post("/api/generate/cancel")
def generate_cancel():
    if ctx.cancel_event:
        ctx.cancel_event.set()
    return {"ok": True}


@app.post("/api/open-output")
def open_output():
    return ctx.open_output_folder()


# ------------------------------------------------------------
# Excel → Excel 搬移
# ------------------------------------------------------------
@app.post("/api/transfer/auto-match")
def transfer_auto_match():
    return ctx.auto_match_transfer_columns()


@app.post("/api/transfer/run")
def transfer_run():
    return ctx.transfer_excel_data()


# ------------------------------------------------------------
# AI 引擎
# ------------------------------------------------------------
@app.get("/api/ai/models")
def ai_models():
    provider = ctx._g("llm_provider", "Gemini")
    try:
        if provider == "Gemini":
            from app.agent.llm import GeminiClient
            c = GeminiClient()
        else:
            from app.agent.llm import OllamaClient
            c = OllamaClient(endpoint=ctx._g("ollama_endpoint"))
        if not c.is_available():
            return {"available": False, "text_models": [], "vision_models": [],
                    "error": f"{provider} 不可用(檢查 API key / endpoint)"}
        return {"available": True,
                "text_models": c.list_models(),
                "vision_models": c.list_vision_models()}
    except Exception as e:
        return {"available": False, "text_models": [], "vision_models": [], "error": str(e)}


@app.get("/api/ai/budget")
def ai_budget():
    return _budget.status()


@app.post("/api/ai/budget/reset")
def ai_budget_reset():
    _budget.reset()
    return _budget.status()


# ------------------------------------------------------------
# Agent 對話(WebSocket)
# ------------------------------------------------------------
@app.websocket("/api/agent/ws")
async def agent_ws(ws: WebSocket):
    await ws.accept()
    import asyncio
    from app.agent.llm import GeminiClient, OllamaClient
    from app.agent.tools import build_default_registry
    from app.agent.orchestrator import AgentOrchestrator

    # 互動工具:透過 websocket 問前端,阻塞等回覆
    pending = {}

    def interaction(kind, payload):
        ev = threading.Event()
        token = str(time.time())
        pending[token] = {"event": ev, "answer": None}
        asyncio.run_coroutine_threadsafe(
            ws.send_json({"type": "interaction", "kind": kind, "token": token, **payload}),
            loop)
        if not ev.wait(timeout=600):
            return {"cancelled": True}
        return pending.pop(token, {}).get("answer") or {"cancelled": True}

    loop = asyncio.get_event_loop()
    agent_ctx = HeadlessContext(settings=ctx.settings, budget=_budget, interaction=interaction)

    provider = agent_ctx._g("llm_provider", "Gemini")
    if provider == "Gemini":
        llm = GeminiClient()
        model = agent_ctx._g("gemini_planner_model")
    else:
        llm = OllamaClient(endpoint=agent_ctx._g("ollama_endpoint"))
        model = agent_ctx._g("ollama_planner_model")

    registry = build_default_registry(agent_ctx)

    if not model:
        await ws.send_json({"type": "error", "text": "尚未選擇 planner 模型,請到「AI 引擎」設定。"})

    orch = AgentOrchestrator(llm=llm, registry=registry, model=model,
                             context=agent_ctx.get_settings(), budget=_budget)

    try:
        while True:
            data = await ws.receive_json()
            mtype = data.get("type")
            if mtype == "interaction_reply":
                token = data.get("token")
                if token in pending:
                    pending[token]["answer"] = data.get("answer")
                    pending[token]["event"].set()
                continue
            if mtype == "reset":
                orch.reset()
                await ws.send_json({"type": "reset_ok"})
                continue
            if mtype == "user":
                text = data.get("text", "")
                orch.add_user_message(text)

                def run_steps():
                    for msg in orch.step():
                        payload = {"type": "message", "role": msg.role,
                                   "text": msg.text, "tool_name": msg.tool_name,
                                   "tool_calls": [{"name": tc.name, "arguments": tc.arguments}
                                                  for tc in msg.tool_calls]}
                        asyncio.run_coroutine_threadsafe(ws.send_json(payload), loop)
                    asyncio.run_coroutine_threadsafe(ws.send_json({"type": "turn_done"}), loop)

                await asyncio.to_thread(run_steps)
    except WebSocketDisconnect:
        pass
    except Exception as e:
        try:
            await ws.send_json({"type": "error", "text": str(e)})
        except Exception:
            pass


# ------------------------------------------------------------
# 渲染預覽(docx → PNG)/ 提供圖片檔
# ------------------------------------------------------------
class RenderBody(BaseModel):
    path: str
    dpi: int = 120
    max_pages: int = 4


@app.post("/api/render/docx")
def render_docx(body: RenderBody):
    return ctx.render_docx_pages(body.path, dpi=body.dpi, max_pages=body.max_pages)


@app.post("/api/render/pptx")
def render_pptx(body: RenderBody):
    return ctx.render_pptx_pages(body.path, max_slides=body.max_pages)


@app.get("/api/file")
def serve_file(path: str):
    if not path or not os.path.isfile(path):
        return {"error": "檔案不存在"}
    return FileResponse(path)


@app.get("/api/health")
def health():
    return {"ok": True}


# ------------------------------------------------------------
# 靜態前端(build 後的 dist)
# ------------------------------------------------------------
_DIST = os.path.join(os.path.dirname(__file__), "static")
if os.path.isdir(_DIST):
    app.mount("/", StaticFiles(directory=_DIST, html=True), name="static")


if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8756)
