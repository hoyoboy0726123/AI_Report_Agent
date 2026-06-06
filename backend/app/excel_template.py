"""Excel 範本標籤模式:讀含 {{tag}} 的 Excel 範本 → 依來源 Excel 每列產出新 Excel。

對照:
- ReportGenerator   → Word 範本 + 來源 Excel → 多份 .docx
- ExcelReportGenerator → Excel 範本 + 來源 Excel → 多份 .xlsx

差異:openpyxl 是「讀取後改 cell.value」,**保留樣式 / 公式 / 列高 / 合併儲存格 / 圖片**;
換值時可保留同 cell 內非 tag 的文字(如 "客戶:{{客戶名稱}}" → "客戶:王小明")。
"""

import os
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from app.config import (
    DEFAULT_HEADER_ROW,
    DEFAULT_OUTPUT_DIR,
    IMAGE_EXTENSIONS,
)
from app.filename import render_filename

DEFAULT_EXCEL_FILENAME_TEMPLATE = "報告_{index}.xlsx"

# 容許 {{ x }} / {{x}} / {{ x.y }};用 [^{}] 防止跨 cell 嚙合
TAG_RE = re.compile(r"\{\{\s*([^{}]+?)\s*\}\}")


class ExcelReportGenerator:
    def __init__(
        self,
        template_path: str,
        excel_path: str,
        output_dir: str = DEFAULT_OUTPUT_DIR,
        sheet_name: str = "",
        header_row: int = DEFAULT_HEADER_ROW,
        filename_template: str = DEFAULT_EXCEL_FILENAME_TEMPLATE,
        image_width_px: int = 320,
    ):
        self.template_path = template_path
        self.excel_path = excel_path
        self.output_dir = output_dir
        self.sheet_name = sheet_name if sheet_name else 0
        self.header_row = max(1, int(header_row or 1))
        self.filename_template = filename_template or DEFAULT_EXCEL_FILENAME_TEMPLATE
        self.image_width_px = max(50, int(image_width_px or 320))

    # ---------- 範本掃描 / 驗證 ----------

    def template_variables(self) -> set:
        """掃整本範本所有 worksheet 中所有 cell 的 {{tag}}。"""
        if not os.path.isfile(self.template_path):
            raise FileNotFoundError(f"Excel 範本不存在: {self.template_path}")
        wb = load_workbook(self.template_path, data_only=False)
        found = set()
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        for m in TAG_RE.findall(cell.value):
                            found.add(m.strip())
        return found

    def list_sheets(self):
        return pd.ExcelFile(self.excel_path).sheet_names

    def _read_dataframe(self):
        return pd.read_excel(
            self.excel_path,
            sheet_name=self.sheet_name,
            header=self.header_row - 1,
        )

    def validate(self):
        """檢查範本變數 vs Excel 欄位;回傳 (missing, extra)。"""
        template_vars = self.template_variables()
        df = self._read_dataframe()
        excel_cols = {str(c) for c in df.columns}
        missing = template_vars - excel_cols
        extra = excel_cols - template_vars
        return missing, extra

    # ---------- 替換邏輯 ----------

    @staticmethod
    def _is_image_path(s):
        return (
            isinstance(s, str)
            and s.lower().endswith(IMAGE_EXTENSIONS)
            and os.path.isfile(s)
        )

    def _resolve_value(self, key: str, row_dict: dict):
        """從 row_dict 取出 key 對應的值,NaN → ""。"""
        if key not in row_dict:
            return None  # 表示找不到 → 保留原 tag
        v = row_dict[key]
        try:
            if pd.isna(v):
                return ""
        except (TypeError, ValueError):
            pass
        return v

    def _render_cell_text(self, text: str, row_dict: dict) -> str:
        """把 cell 文字內所有 {{tag}} 換成對應值;找不到的 tag 保留原樣。"""
        def replacer(m):
            key = m.group(1).strip()
            v = self._resolve_value(key, row_dict)
            if v is None:
                return m.group(0)  # 保留原 tag
            return str(v)
        return TAG_RE.sub(replacer, text)

    def _maybe_image_replace(self, ws, cell, row_dict):
        """若 cell 完整內容只是單一 {{tag}} 且其值是圖片路徑 → 插入圖片並清空 cell。
        回傳 True 表示已處理為圖片,呼叫端應跳過普通文字替換。
        """
        text = cell.value
        if not isinstance(text, str):
            return False
        m = TAG_RE.fullmatch(text.strip())
        if not m:
            return False
        key = m.group(1).strip()
        v = self._resolve_value(key, row_dict)
        if not self._is_image_path(v):
            return False
        try:
            img = XLImage(v)
            # 用像素寬縮放(高保持比例)
            ratio = self.image_width_px / float(img.width or self.image_width_px)
            img.width = self.image_width_px
            img.height = int((img.height or self.image_width_px) * ratio)
            img.anchor = cell.coordinate
            ws.add_image(img)
            cell.value = None
            return True
        except Exception:
            # 失敗就退回字串模式
            return False

    # ---------- 產出 ----------

    def generate_iter(self, cancel_event=None):
        """逐列產出 Excel,yield (produced, total, saved_path, row_dict)。"""
        df = self._read_dataframe()
        os.makedirs(self.output_dir, exist_ok=True)

        total = len(df)
        produced = 0
        for index, row in df.iterrows():
            if cancel_event is not None and cancel_event.is_set():
                return

            wb = load_workbook(self.template_path)
            row_dict = row.to_dict()

            for ws in wb.worksheets:
                # 收集所有要改的 cell 先(避免迭代中改值的副作用)
                targets = []
                for r in ws.iter_rows():
                    for cell in r:
                        if isinstance(cell.value, str) and TAG_RE.search(cell.value):
                            targets.append(cell)
                for cell in targets:
                    if self._maybe_image_replace(ws, cell, row_dict):
                        continue
                    cell.value = self._render_cell_text(cell.value, row_dict)

            filename = render_filename(
                self.filename_template, row_dict, index + 1, default_ext=".xlsx"
            )
            saved_path = os.path.join(self.output_dir, filename)
            wb.save(saved_path)

            produced += 1
            yield produced, total, saved_path, row_dict

    def generate(self, progress_callback=None, cancel_event=None):
        produced = 0
        total = 0
        for prod, tot, _, _ in self.generate_iter(cancel_event=cancel_event):
            produced = prod
            total = tot
            if progress_callback:
                progress_callback(produced, total)
        return produced, total


# ============================================================
# 範本共用靜態圖(Excel):把值=單一 {{欄位}} 的儲存格就地換成圖片
# ============================================================

def replace_tag_with_image_xlsx(xlsx_path, field, image_path, width_px=320, output_path=None):
    """找儲存格內容整個是 {{field}} 的格,插圖並清空。回 {replaced}。"""
    if not os.path.isfile(xlsx_path):
        return {"error": f"Excel 不存在: {xlsx_path}"}
    if not image_path or not os.path.isfile(image_path):
        return {"error": f"圖片不存在: {image_path}"}
    pat = re.compile(r"\{\{\s*" + re.escape(field) + r"\s*\}\}")
    wb = load_workbook(xlsx_path)
    replaced = 0
    for ws in wb.worksheets:
        targets = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and pat.fullmatch(cell.value.strip()):
                    targets.append(cell)
        for cell in targets:
            try:
                img = XLImage(image_path)
                ratio = width_px / float(img.width or width_px)
                img.width = width_px
                img.height = int((img.height or width_px) * ratio)
                img.anchor = cell.coordinate
                ws.add_image(img)
                cell.value = None
                replaced += 1
            except Exception as e:
                return {"error": f"插入圖片失敗: {e}"}
    if replaced:
        wb.save(output_path or xlsx_path)
    return {"replaced": replaced, "field": field}


# ============================================================
# 無標註文字填寫(Excel 表單):把一列資料依語意填到標籤格旁
# ============================================================

def auto_fill_text_xlsx(llm, model, xlsx_path, data, output_path):
    """Excel 表單無 {{標籤}} 時,讓 LLM 判斷每個欄位值該填到哪個儲存格(標籤右/下/同格後)。"""
    import json as _json
    from openpyxl.utils import coordinate_to_tuple, get_column_letter
    from app.agent.llm.base import Message
    if not os.path.isfile(xlsx_path):
        return {"error": f"Excel 不存在: {xlsx_path}"}
    data = {str(k): ("" if v is None else str(v)) for k, v in (data or {}).items()}
    data = {k: v for k, v in data.items() if v.strip()}
    if not data:
        return {"error": "沒有可填的資料"}

    wb = load_workbook(xlsx_path)
    cells = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.strip():
                    cells.append({"sheet": ws.title, "addr": c.coordinate, "text": c.value.strip()[:50]})
    if not cells:
        return {"error": "表單沒有可定位的標籤文字"}

    sys = ("你是表單填寫助理。會給你『欄位:值』資料,與 Excel 既有的文字儲存格(每個有 sheet、addr、text)。"
           "請判斷每個欄位的值該填在哪 —— 通常標籤在某格(如「客戶」),值要填到它『右邊一格』(mode=right);"
           "若標籤同格已含冒號(如「客戶:」)則 mode=after(接在同格後);值在標籤下方則 mode=below。"
           "只回 JSON:{\"fills\":[{\"field\":\"欄位\",\"sheet\":\"工作表\",\"addr\":\"標籤格座標\",\"mode\":\"right|below|after\"}]};對不到略過。")
    user = _json.dumps({"資料": data, "標籤格": cells}, ensure_ascii=False)
    try:
        resp = llm.chat([Message(role="system", text=sys), Message(role="user", text=user)],
                        model=model, tools=None)
        t = (resp.text or "").strip()
        if t.startswith("```"):
            t = re.sub(r"^```\w*\n?", "", t); t = re.sub(r"\n?```\s*$", "", t)
        fills = _json.loads(t).get("fills", [])
    except Exception as e:
        return {"error": f"LLM 配對失敗: {e}"}

    filled, failed = [], []
    for f in fills:
        field = f.get("field"); sheet = f.get("sheet"); addr = f.get("addr"); mode = f.get("mode", "right")
        if field not in data:
            continue
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        try:
            r, col = coordinate_to_tuple(addr)
            val = data[field]
            if mode == "after":
                ws[addr] = f"{ws[addr].value or ''}{val}"; tgt = addr
            elif mode == "below":
                ws.cell(row=r + 1, column=col, value=val); tgt = f"{get_column_letter(col)}{r+1}"
            else:
                ws.cell(row=r, column=col + 1, value=val); tgt = f"{get_column_letter(col+1)}{r}"
            filled.append({"field": field, "cell": tgt, "mode": mode})
        except Exception as e:
            failed.append({"field": field, "reason": str(e)})

    if not output_path:
        b, ext = os.path.splitext(xlsx_path)
        output_path = f"{b}_filled{ext}"
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return {"ok": True, "output": output_path, "filled": filled, "fill_failed": failed,
            "filled_count": len(filled), "total_fields": len(data)}


# ---------- Excel cell 端的 hotkey 寫入(非批次,即時編輯範本) ----------

def insert_tag_into_excel_cell(template_path: str, sheet_name: str, cell_addr: str, tag: str):
    """把 {{ tag }} 寫進 template_path 中指定 sheet 的 cell;若 cell 已有文字則前面追加。
    純檔案操作,不需 Excel 開啟;呼叫端負責提供 sheet + cell address。
    """
    if not os.path.isfile(template_path):
        raise FileNotFoundError(f"Excel 範本不存在: {template_path}")
    if not tag:
        raise ValueError("tag 不可為空。")
    wb = load_workbook(template_path)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    cell = ws[cell_addr]
    placeholder = f"{{{{ {tag} }}}}"
    existing = cell.value
    if existing is None or str(existing).strip() == "":
        cell.value = placeholder
    else:
        # 既有內容前面接,避免破壞原 label
        cell.value = f"{existing}{placeholder}"
    wb.save(template_path)
    return cell.coordinate
