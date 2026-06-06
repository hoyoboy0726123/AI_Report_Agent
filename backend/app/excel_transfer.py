"""Excel → Excel 資料搬移：依欄位對應將來源 Excel 的列複製到目標 Excel。

設計：
- 來源用 pandas 讀（可指定 sheet / 標題列）
- 目標用 openpyxl 操作（保留既有格式 / 樣式 / 公式）
- column_map: {來源欄位: 目標欄位}
- mode:
    - "append"     資料寫到目標既有資料下方（找第一個空白列起）
    - "overwrite"  清空標題列以下所有儲存格再寫入
    - "fresh"      新建目標檔（或新建目標 sheet），自動寫標題列
"""

import os

import pandas as pd
from openpyxl import Workbook, load_workbook


VALID_MODES = ("append", "overwrite", "fresh")


class ExcelTransfer:
    def __init__(
        self,
        source_path: str,
        target_path: str,
        column_map: dict,
        source_sheet: str = "",
        source_header_row: int = 1,
        target_sheet: str = "",
        target_header_row: int = 1,
        mode: str = "append",
    ):
        self.source_path = source_path
        self.target_path = target_path
        self.column_map = {
            str(k): str(v) for k, v in (column_map or {}).items() if v
        }
        self.source_sheet = source_sheet or 0
        self.source_header_row = max(1, int(source_header_row or 1))
        self.target_sheet = target_sheet or ""
        self.target_header_row = max(1, int(target_header_row or 1))
        self.mode = mode if mode in VALID_MODES else "append"

    # ---------- helpers ----------

    @staticmethod
    def list_sheets(path: str):
        if not path or not os.path.isfile(path):
            return []
        try:
            return pd.ExcelFile(path).sheet_names
        except Exception:
            return []

    @staticmethod
    def list_columns(path: str, sheet: str = "", header_row: int = 1):
        if not path or not os.path.isfile(path):
            return []
        try:
            df = pd.read_excel(
                path,
                sheet_name=sheet if sheet else 0,
                header=max(0, int(header_row) - 1),
                nrows=0,
            )
            return [str(c) for c in df.columns]
        except Exception:
            return []

    def _read_source(self):
        return pd.read_excel(
            self.source_path,
            sheet_name=self.source_sheet,
            header=self.source_header_row - 1,
        )

    def _open_target(self):
        if self.mode == "fresh" or not os.path.isfile(self.target_path):
            wb = Workbook()
            # 預設活頁簿會帶一個 "Sheet"；若指定了 sheet 名稱則改名
            default_ws = wb.active
            default_ws.title = self.target_sheet or "Sheet1"
            return wb, True
        return load_workbook(self.target_path), False

    # ---------- main ----------

    def transfer(self, progress_callback=None, cancel_event=None):
        """執行搬移；回傳 dict {rows_written, total, target_path, sheet, mode, cancelled}。"""
        if not os.path.isfile(self.source_path):
            raise FileNotFoundError(f"來源 Excel 不存在：{self.source_path}")
        if not self.column_map:
            raise ValueError("欄位對應為空，請先指定 來源欄位 → 目標欄位。")
        if not self.target_path:
            raise ValueError("未指定目標 Excel 路徑。")

        df = self._read_source()
        wb, created_new = self._open_target()

        sheet_name = self.target_sheet or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

        header_row = self.target_header_row
        target_cols = sorted(set(self.column_map.values()))

        # 取得 / 補齊目標標題列 → {col_name: col_index}
        target_header_map = {}
        if self.mode != "fresh" and not created_new:
            for col_idx in range(1, max(ws.max_column, 1) + 1):
                v = ws.cell(row=header_row, column=col_idx).value
                if v is not None and str(v).strip():
                    target_header_map[str(v)] = col_idx

        next_col = (max(target_header_map.values()) if target_header_map else 0) + 1
        for tgt_col in target_cols:
            if tgt_col not in target_header_map:
                ws.cell(row=header_row, column=next_col, value=tgt_col)
                target_header_map[tgt_col] = next_col
                next_col += 1

        # 起始寫入列
        if self.mode == "overwrite":
            max_row = ws.max_row or header_row
            for r in range(header_row + 1, max_row + 1):
                for c in range(1, (ws.max_column or 1) + 1):
                    ws.cell(row=r, column=c).value = None
            start_row = header_row + 1
        elif self.mode == "fresh":
            start_row = header_row + 1
        else:  # append
            r = header_row + 1
            max_row = ws.max_row or header_row
            while r <= max_row:
                if all(
                    ws.cell(row=r, column=c).value in (None, "")
                    for c in range(1, (ws.max_column or 1) + 1)
                ):
                    break
                r += 1
            start_row = r

        # 寫入資料
        rows_written = 0
        total = len(df)
        src_cols_set = set(df.columns.astype(str))
        for i, (_, src_row) in enumerate(df.iterrows()):
            if cancel_event is not None and cancel_event.is_set():
                break
            for src_col, tgt_col in self.column_map.items():
                if src_col not in src_cols_set:
                    continue
                v = src_row[src_col]
                try:
                    if pd.isna(v):
                        v = None
                except (TypeError, ValueError):
                    pass
                ws.cell(
                    row=start_row + i,
                    column=target_header_map[tgt_col],
                    value=v,
                )
            rows_written += 1
            if progress_callback:
                progress_callback(rows_written, total)

        # 確保目標資料夾存在
        target_dir = os.path.dirname(self.target_path)
        if target_dir:
            os.makedirs(target_dir, exist_ok=True)
        wb.save(self.target_path)

        return {
            "rows_written": rows_written,
            "total": total,
            "target_path": self.target_path,
            "sheet": sheet_name,
            "mode": self.mode,
            "cancelled": cancel_event is not None and cancel_event.is_set(),
        }
