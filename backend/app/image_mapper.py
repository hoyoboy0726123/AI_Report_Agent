"""資料夾圖片 → 欄位 對應。

兩種情境:
- 靜態範本欄位(static):圖片依檔名對到 Word 範本的 {{欄位}},產出前一次插好,所有報告共用。
- 每列不同(per-row):圖片依檔名對到「來源 Excel 某 key 欄位的值」,把圖片路徑填進指定圖片欄,
  再由既有 generator 在產出時逐列嵌入(Excel 儲存格=圖片路徑的機制)。

對應方式:
- 確定性比對(檔名 stem 對 target,精確 / 正規化 / 包含)
- AI 比對(用 planner LLM 依檔名語意挑最適欄位)
"""

import json
import os
import re

import pandas as pd
from openpyxl import load_workbook

from app.config import IMAGE_EXTENSIONS


def _stem(name: str) -> str:
    return os.path.splitext(os.path.basename(name))[0]


def _norm(s: str) -> str:
    """正規化:小寫、去空白/底線/連字號,便於模糊比對。"""
    return re.sub(r"[\s_\-]+", "", str(s).strip().lower())


def list_images(folder: str) -> list:
    """回傳 [{name, path}]。"""
    if not folder or not os.path.isdir(folder):
        return []
    out = []
    for fn in sorted(os.listdir(folder)):
        full = os.path.join(folder, fn)
        if os.path.isfile(full) and fn.lower().endswith(IMAGE_EXTENSIONS):
            out.append({"name": fn, "path": full})
    return out


def deterministic_match(image_names: list, targets: list) -> dict:
    """以檔名比對 targets,回傳 {image_name: target 或 ""}。

    優先序:正規化精確 > target 包含於檔名 > 檔名包含於 target。
    """
    norm_targets = [(t, _norm(t)) for t in targets]
    result = {}
    for img in image_names:
        s = _norm(_stem(img))
        chosen = ""
        # 1. 精確
        for t, nt in norm_targets:
            if nt and nt == s:
                chosen = t
                break
        # 2. target 出現在檔名中(如 "宏達科技_店面" 含 "店面")
        if not chosen:
            cands = [t for t, nt in norm_targets if nt and nt in s]
            if cands:
                chosen = max(cands, key=lambda t: len(_norm(t)))  # 取最長(最具體)
        # 3. 檔名出現在 target 中
        if not chosen:
            for t, nt in norm_targets:
                if s and s in nt:
                    chosen = t
                    break
        result[img] = chosen
    return result


def ai_match(llm, model: str, image_names: list, targets: list, hint: str = "") -> dict:
    """用 LLM 依檔名語意把每張圖配到最適 target;回傳 {image_name: target 或 ""}。"""
    sys = (
        "你是檔案對應助理。使用者有一批圖片檔名,要對應到一組欄位名稱。"
        "依檔名語意判斷每張圖最該放到哪個欄位。"
        "只能從提供的欄位清單挑選,真的對不上就回空字串。"
        "務必只回傳 JSON 物件,key 是圖片檔名,value 是欄位名稱(或空字串),不要任何解釋。"
    )
    user = json.dumps({
        "圖片檔名": image_names,
        "可選欄位": targets,
        "情境提示": hint or "",
    }, ensure_ascii=False)

    from app.agent.llm.base import Message
    try:
        resp = llm.chat([Message(role="system", text=sys), Message(role="user", text=user)],
                        model=model, tools=None)
        txt = (resp.text or "").strip()
        if txt.startswith("```"):
            txt = re.sub(r"^```\w*\n?", "", txt)
            txt = re.sub(r"\n?```\s*$", "", txt)
        data = json.loads(txt)
        if not isinstance(data, dict):
            return {"error": "AI 回傳格式非物件"}
        tset = set(targets)
        return {img: (data.get(img) if data.get(img) in tset else "") for img in image_names}
    except Exception as e:
        return {"error": f"AI 對應失敗: {e}"}


def apply_static_to_template(word_path: str, mapping: dict, width_mm: int = 80) -> dict:
    """mapping = {欄位名: 圖片路徑};把範本每個 {{欄位}} 換成圖片。依副檔名支援 docx/pptx/xlsx。"""
    if not word_path or not os.path.isfile(word_path):
        return {"error": f"範本不存在: {word_path}"}
    low = word_path.lower()

    def _docx(field, img):
        from app.agent.template_edit import replace_tag_with_image
        return replace_tag_with_image(word_path, field, img, width_mm=width_mm)

    def _pptx(field, img):
        from app.pptx_template import replace_tag_with_image_pptx
        return replace_tag_with_image_pptx(word_path, field, img)

    def _xlsx(field, img):
        from app.excel_template import replace_tag_with_image_xlsx
        return replace_tag_with_image_xlsx(word_path, field, img, width_px=max(120, int(width_mm * 3.78)))

    fn = _pptx if low.endswith(".pptx") else _xlsx if low.endswith((".xlsx", ".xls")) else _docx

    applied, failed = [], []
    for field, img in mapping.items():
        if not img:
            continue
        r = fn(field, img)
        if r.get("error") or r.get("replaced", 0) == 0:
            failed.append({"field": field, "image": img, "reason": r.get("error") or r.get("warning")})
        else:
            applied.append({"field": field, "image": os.path.basename(img)})
    return {"applied": applied, "failed": failed, "applied_count": len(applied)}


def fill_excel_image_column(
    excel_path: str, sheet: str, header_row: int,
    key_column: str, image_column: str, folder: str,
    image_to_key: dict = None,
) -> dict:
    """為每一列找對應圖片,把路徑寫進 image_column,另存副本 *_with_images.xlsx。

    image_to_key:可選 {圖片檔名: key 值}(AI/手動覆寫);未提供則用檔名 stem 對 key 值。
    回傳 {output_path, matched, total, unmatched_rows}。
    """
    if not excel_path or not os.path.isfile(excel_path):
        return {"error": f"來源 Excel 不存在: {excel_path}"}
    images = list_images(folder)
    if not images:
        return {"error": f"資料夾無圖片: {folder}"}

    # 建立 key 值 → 圖片路徑
    df = pd.read_excel(excel_path, sheet_name=sheet if sheet else 0, header=max(0, header_row - 1))
    if key_column not in [str(c) for c in df.columns]:
        return {"error": f"來源沒有 key 欄位: {key_column}"}
    key_values = [str(v) for v in df[key_column].tolist()]

    if image_to_key:
        key_to_path = {}
        name_to_path = {im["name"]: im["path"] for im in images}
        for img_name, keyval in image_to_key.items():
            if keyval and img_name in name_to_path:
                key_to_path[_norm(keyval)] = name_to_path[img_name]
    else:
        # 檔名 stem 對 key 值
        match = deterministic_match([im["name"] for im in images], key_values)
        name_to_path = {im["name"]: im["path"] for im in images}
        key_to_path = {}
        for img_name, keyval in match.items():
            if keyval:
                key_to_path[_norm(keyval)] = name_to_path[img_name]

    # 寫進 openpyxl(保留樣式)
    wb = load_workbook(excel_path)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    # 找/建 image_column 的欄位索引
    col_idx = None
    for c in range(1, (ws.max_column or 1) + 1):
        if str(ws.cell(row=header_row, column=c).value) == image_column:
            col_idx = c
            break
    if col_idx is None:
        col_idx = (ws.max_column or 0) + 1
        ws.cell(row=header_row, column=col_idx, value=image_column)

    matched = 0
    unmatched = []
    for i, kv in enumerate(key_values):
        path = key_to_path.get(_norm(kv))
        if path:
            ws.cell(row=header_row + 1 + i, column=col_idx, value=path)
            matched += 1
        else:
            unmatched.append(kv)

    base, ext = os.path.splitext(excel_path)
    out_path = f"{base}_with_images{ext or '.xlsx'}"
    wb.save(out_path)
    return {
        "output_path": out_path,
        "matched": matched,
        "total": len(key_values),
        "unmatched_rows": unmatched[:20],
        "image_column": image_column,
    }
